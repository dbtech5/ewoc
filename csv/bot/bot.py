from sqlite3 import Row
import pandas as pd
import openpyxl
from pyparsing import col
import json
import os
dir = os.listdir('file_xlsx')
print(dir)
for item in dir:
    item = item.replace('~$','')
    
    if True:
        file_name = './file_xlsx/'+item 

        book = openpyxl.load_workbook(file_name, data_only = True)
        sheet_name = book.sheetnames
        name_file = './file_json/อ่างเก็บน้ำ'+sheet_name[1]+'.json'
        json_format = {}
        sheet = book[sheet_name[3]]
        rulecurve = None
        try:
            rulecurve = book['rulecurve']
        except:
            rulecurve = book['ruleCurve']
        #print(name_file)
        #print(sheet_name[3])
        #print(sheet.cell(row=5,column = 2).value)
        pattern_month = {
            '01':'Jan',
            '02':'Feb',
            '03':'Mar',
            '04':'Apr',
            '05':'May',
            '06':'Jun',
            '07':'Jul',
            '08':'Aug',
            '09':'Sep',
            '10':'Oct',
            '11':'Nov',
            '12':'Dec',    
        }

        rulecurve_rows = 2
        r = 5
        c = 2
        data_feed = str(sheet.cell(row=5,column = 2).value)
        #print(sheet_name[3])
        while(data_feed != str(None)):
            data_feed = data_feed.split('-')
            tmp_text = data_feed[2].split(' ')[0] +' '+ pattern_month[data_feed[1]]
            json_inner = {}
            json_inner[sheet.cell(row=4,column = 3).value] = sheet.cell(row=r,column = 3).value
            json_inner[sheet.cell(row=4,column = 4).value] = sheet.cell(row=r,column = 4).value
            json_inner[sheet.cell(row=4,column = 5).value] = sheet.cell(row=r,column = 5).value
            json_inner[sheet.cell(row=4,column = 6).value] = sheet.cell(row=r,column = 6).value
            json_inner[sheet.cell(row=4,column = 7).value] = sheet.cell(row=r,column = 7).value
            json_inner['Max'] = sheet.cell(row=r,column = 8).value
            json_inner['NHWL'] = sheet.cell(row=r,column = 9).value
            json_inner['Min'] = sheet.cell(row=r,column = 10).value
            
            cl_in = 15
            cl_out = 16
            if(sheet.cell(row=3,column = 15).value == "Outflow"):
                cl_in = 14
                cl_out = 15
            json_inner['inflow'] = sheet.cell(row=r,column = cl_in).value
            json_inner['outflow'] = sheet.cell(row=r,column = cl_out).value
            if(data_feed[2].split(' ')[0]=='01'):
                json_inner['upper_rc'] = rulecurve.cell(row=rulecurve_rows,column=2).value
                json_inner['lower_rc'] = rulecurve.cell(row=rulecurve_rows,column=3).value
                json_inner['mrc_rc'] = rulecurve.cell(row=rulecurve_rows,column=5).value
                rulecurve_rows += 1    
            #print(data_feed[2].split(' ')[0],rulecurve.cell(row=rulecurve_rows,column=2).value)
            
            json_format[tmp_text] = json_inner
            r+=1
            data_feed = str(sheet.cell(row=r,column = 2).value)


        json_object = json.dumps(json_format)

        with open(name_file,'w') as out:
            out.write(json_object)
        print('write file json '+item)
        
    else:
        print('error filename '+item)
    